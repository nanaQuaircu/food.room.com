#!/usr/bin/env python3
"""Import Hotel Messiah warehouse Excel into hotel_grand_plaza_hotel."""
from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
import pymysql

EXCEL_PATH = Path(
    r"c:\Users\Duapa Werkspace 007\Downloads\Hotel_Messiah_Warehouse_JUNE finale (1).xlsx"
)
DB = dict(host="127.0.0.1", user="root", password="", database="hotel_grand_plaza_hotel", charset="utf8mb4")
PROPERTY_ID = 1
LOC = {"warehouse": 1, "kitchen": 2, "cleaners": 3, "front_office": 4}

issues: list[str] = []


def log(msg: str) -> None:
    print(msg, flush=True)


def norm_key(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip()).lower()


def to_decimal(val, default=Decimal("0")) -> Decimal:
    if val is None or val == "":
        return default
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    if isinstance(val, datetime):
        return default
    try:
        return Decimal(str(val).strip().replace(",", ""))
    except (InvalidOperation, ValueError):
        return default


def to_date(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def normalize_dept(dept) -> str:
    if dept is None or str(dept).strip() == "":
        return "general"
    d = norm_key(dept)
    if "clean" in d:
        return "cleaners"
    if "front" in d and "office" in d:
        return "front_office"
    if "kitchen" in d or "restaurant" in d or d == "food":
        return "kitchen"
    if "warehouse" in d:
        return "warehouse"
    return re.sub(r"[^a-z0-9]+", "_", d).strip("_") or "general"


def normalize_dest(dest) -> str | None:
    if dest is None:
        return None
    d = norm_key(dest)
    if "kitchen" in d:
        return "kitchen"
    if "clean" in d:
        return "cleaners"
    if "front" in d and "office" in d:
        return "front_office"
    if "warehouse" in d:
        return "warehouse"
    return None


def find_header_row(ws, marker: str) -> int | None:
    marker_l = marker.lower()
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] and marker_l in str(row[0]).lower():
            return idx
    return None


def sheet_rows_after_header(ws, header_marker: str):
    started = False
    for row in ws.iter_rows(values_only=True):
        if not started:
            if row and row[0] and header_marker.lower() in str(row[0]).lower():
                started = True
            continue
        yield row


def clear_property_stock(cur) -> None:
    pid = PROPERTY_ID
    cur.execute(
        "DELETE FROM stock_usage_logs WHERE property_id = %s",
        (pid,),
    )
    cur.execute(
        """
        DELETE stl FROM stock_transfer_lines stl
        INNER JOIN stock_transfers st ON st.id = stl.transfer_id
        WHERE st.property_id = %s
        """,
        (pid,),
    )
    cur.execute("DELETE FROM stock_transfers WHERE property_id = %s", (pid,))
    cur.execute(
        """
        DELETE spl FROM stock_purchase_lines spl
        INNER JOIN stock_purchases sp ON sp.id = spl.purchase_id
        WHERE sp.property_id = %s
        """,
        (pid,),
    )
    cur.execute("DELETE FROM stock_purchases WHERE property_id = %s", (pid,))
    cur.execute(
        """
        DELETE sb FROM stock_balances sb
        INNER JOIN stock_items si ON si.id = sb.item_id
        WHERE si.property_id = %s
        """,
        (pid,),
    )
    cur.execute("DELETE FROM stock_unit_conversions WHERE property_id = %s", (pid,))
    cur.execute("DELETE FROM stock_items WHERE property_id = %s", (pid,))


def build_item_lookup(items: dict[int, dict]) -> tuple[dict, dict]:
    by_sku: dict[str, int] = {}
    by_name: dict[str, int] = {}
    for iid, row in items.items():
        if row.get("sku"):
            by_sku[norm_key(row["sku"])] = iid
        by_name[norm_key(row["name"])] = iid
    return by_sku, by_name


def resolve_item(by_sku, by_name, code, name) -> int | None:
    if code and norm_key(code) in by_sku:
        return by_sku[norm_key(code)]
    if name and norm_key(name) in by_name:
        return by_name[norm_key(name)]
    # fuzzy: name contains
    nk = norm_key(name)
    if nk:
        for k, iid in by_name.items():
            if k == nk or k.startswith(nk) or nk.startswith(k):
                return iid
    return None


def import_master(wb, cur) -> dict[int, dict]:
    ws = wb["Warehouse Inventory"]
    items: dict[int, dict] = {}
    warehouse_qty: dict[int, Decimal] = {}

    for row in sheet_rows_after_header(ws, "Item Code"):
        code = row[0]
        if code is None or str(code).strip() == "":
            continue
        name = row[1]
        if name is None or str(name).strip() == "":
            issues.append(f"Skip row with code {code}: missing name")
            continue
        category = row[2]
        dept = normalize_dept(row[3])
        unit = (str(row[4]).strip() if row[4] else "unit") or "unit"
        unit_cost = to_decimal(row[5])
        initial = to_decimal(row[6])
        current = to_decimal(row[10]) if len(row) > 10 else Decimal("0")
        reorder = to_decimal(row[12]) if len(row) > 12 else Decimal("0")
        qoh = current if current > 0 else initial

        cur.execute(
            """
            INSERT INTO stock_items
              (property_id, name, sku, department, category, unit, quantity_on_hand, reorder_level, unit_cost)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                PROPERTY_ID,
                str(name).strip(),
                str(code).strip(),
                dept,
                str(category).strip() if category else None,
                unit[:20],
                qoh,
                reorder,
                unit_cost,
            ),
        )
        iid = cur.lastrowid
        items[iid] = {
            "name": str(name).strip(),
            "sku": str(code).strip(),
            "current_master": current,
            "initial": initial,
        }
        wh_q = current if current > 0 else initial
        warehouse_qty[iid] = wh_q
        cur.execute(
            """
            INSERT INTO stock_balances (item_id, location_id, quantity)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE quantity = VALUES(quantity)
            """,
            (iid, LOC["warehouse"], wh_q),
        )

    return items, warehouse_qty




def ensure_items_from_location_sheet(ws, cur, by_sku, by_name, dept_default: str) -> None:
    for row in sheet_rows_after_header(ws, "Item Code"):
        code, name = row[0], row[1]
        if not code or str(code).strip().upper() == "TOTALS":
            continue
        if not name or str(name).strip() == "":
            continue
        if resolve_item(by_sku, by_name, code, name):
            continue
        unit = (str(row[3]).strip() if len(row) > 3 and row[3] else "unit") or "unit"
        unit_cost = to_decimal(row[4]) if len(row) > 4 else Decimal("0")
        reorder = to_decimal(row[8]) if len(row) > 8 else Decimal("0")
        cur.execute(
            """
            INSERT INTO stock_items
              (property_id, name, sku, department, category, unit, quantity_on_hand, reorder_level, unit_cost)
            VALUES (%s, %s, %s, %s, %s, %s, 0, %s, %s)
            """,
            (
                PROPERTY_ID,
                str(name).strip(),
                str(code).strip(),
                dept_default,
                None,
                unit[:20],
                reorder,
                unit_cost,
            ),
        )
        iid = cur.lastrowid
        by_sku[norm_key(code)] = iid
        by_name[norm_key(name)] = iid

def apply_location_sheet(ws, cur, by_sku, by_name, loc_id: int, qty_col: int, label: str) -> None:
    for row in sheet_rows_after_header(ws, "Item Code"):
        code, name = row[0], row[1]
        if not code and not name:
            continue
        if code and str(code).strip().upper() == "TOTALS":
            continue
        if code and str(code).strip().lower().startswith("item"):
            continue
        iid = resolve_item(by_sku, by_name, code, name)
        if not iid:
            issues.append(f"{label}: no item match for {code!r} / {name!r}")
            continue
        qty = to_decimal(row[qty_col]) if len(row) > qty_col else Decimal("0")
        cur.execute(
            """
            INSERT INTO stock_balances (item_id, location_id, quantity)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE quantity = VALUES(quantity)
            """,
            (iid, loc_id, qty),
        )


def sync_quantity_on_hand(cur, items: dict[int, dict]) -> None:
    for iid, meta in items.items():
        cur.execute(
            "SELECT COALESCE(SUM(quantity), 0) FROM stock_balances WHERE item_id = %s",
            (iid,),
        )
        total = Decimal(str(cur.fetchone()[0]))
        current_master = meta.get("current_master") or Decimal("0")
        if total <= 0 and current_master > 0:
            qoh = current_master
        else:
            qoh = total
        cur.execute("UPDATE stock_items SET quantity_on_hand = %s WHERE id = %s", (qoh, iid))


def import_purchases(ws, cur, by_sku, by_name) -> tuple[int, int]:
    groups: dict[str, list] = defaultdict(list)
    for row in sheet_rows_after_header(ws, "Purchase ID"):
        ref = row[0]
        if ref is None or str(ref).strip() == "":
            continue
        ref_s = str(ref).strip().upper()
        if not ref_s.startswith("PUR"):
            continue
        groups[ref_s].append(row)

    n_purchases = n_lines = 0
    for ref, lines in sorted(groups.items()):
        try:
            first = lines[0]
            pdate = to_date(first[1]) or date.today()
            total = Decimal("0")
            line_rows = []
            for row in lines:
                item_name = row[2]
                qty = to_decimal(row[3])
                unit_cost = to_decimal(row[4])
                line_total = to_decimal(row[5])
                if line_total == 0 and qty and unit_cost:
                    line_total = qty * unit_cost
                iid = resolve_item(by_sku, by_name, None, item_name)
                if not iid:
                    issues.append(f"Purchase {ref}: unknown item {item_name!r}")
                    continue
                total += line_total
                line_rows.append((iid, qty, unit_cost, line_total))
            if not line_rows:
                continue
            notes = None
            supplier = first[6] if len(first) > 6 else None
            if supplier:
                notes = f"Supplier: {supplier}"
            cur.execute(
                """
                INSERT INTO stock_purchases
                  (property_id, location_id, reference, purchase_date, total_amount, notes)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (PROPERTY_ID, LOC["warehouse"], ref, pdate, total, notes),
            )
            pid = cur.lastrowid
            n_purchases += 1
            for iid, qty, unit_cost, line_total in line_rows:
                cur.execute(
                    """
                    INSERT INTO stock_purchase_lines
                      (purchase_id, item_id, quantity, unit_cost, line_total)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (pid, iid, qty, unit_cost, line_total),
                )
                n_lines += 1
        except Exception as e:
            issues.append(f"Purchase group {ref} skipped: {e}")
    return n_purchases, n_lines


def import_transfers(ws, cur, by_sku, by_name) -> tuple[int, int]:
    groups: dict[str, list] = defaultdict(list)
    for row in sheet_rows_after_header(ws, "Transfer ID"):
        ref = row[0]
        if ref is None or str(ref).strip() == "":
            continue
        ref_s = str(ref).strip().upper()
        if not ref_s.startswith("TRF"):
            continue
        groups[ref_s].append(row)

    n_transfers = n_lines = 0
    for ref, lines in sorted(groups.items()):
        try:
            first = lines[0]
            tdate = to_date(first[1]) or date.today()
            dest_code = normalize_dest(first[10] if len(first) > 10 else None)
            if not dest_code or dest_code == "warehouse":
                dest_code = normalize_dest(first[9] if len(first) > 9 else None)
            if not dest_code or dest_code not in LOC or dest_code == "warehouse":
                issues.append(f"Transfer {ref}: bad destination {first[10]!r}")
                continue
            to_loc = LOC[dest_code]
            line_rows = []
            for row in lines:
                item_name = row[2]
                qty = to_decimal(row[5]) if len(row) > 5 else Decimal("0")
                if qty <= 0:
                    continue
                iid = resolve_item(by_sku, by_name, None, item_name)
                if not iid:
                    issues.append(f"Transfer {ref}: unknown item {item_name!r}")
                    continue
                line_rows.append((iid, qty))
            if not line_rows:
                continue
            notes = str(first[13]).strip() if len(first) > 13 and first[13] else None
            cur.execute(
                """
                INSERT INTO stock_transfers
                  (property_id, from_location_id, to_location_id, reference, transfer_date, notes)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (PROPERTY_ID, LOC["warehouse"], to_loc, ref, tdate, notes),
            )
            tid = cur.lastrowid
            n_transfers += 1
            for iid, qty in line_rows:
                cur.execute(
                    """
                    INSERT INTO stock_transfer_lines (transfer_id, item_id, quantity)
                    VALUES (%s, %s, %s)
                    """,
                    (tid, iid, qty),
                )
                n_lines += 1
        except Exception as e:
            issues.append(f"Transfer group {ref} skipped: {e}")
    return n_transfers, n_lines


def import_conversions(wb, cur, by_sku, by_name) -> int:
    ws = wb["Conversions"]
    seen: set[tuple] = set()
    count = 0
    in_section2 = False
    for row in ws.iter_rows(values_only=True):
        c0 = row[0] if row else None
        if c0 and "SECTION 2" in str(c0).upper():
            in_section2 = True
            continue
        if in_section2:
            if c0 and str(c0).strip().upper().startswith("PUR"):
                item_name = row[2] if len(row) > 2 else None
                pieces_per = to_decimal(row[5]) if len(row) > 5 else Decimal("0")
                if item_name and pieces_per > 0:
                    iid = resolve_item(by_sku, by_name, None, item_name)
                    if iid:
                        key = (iid, "box", "pieces")
                        if key not in seen:
                            seen.add(key)
                            cur.execute(
                                """
                                INSERT INTO stock_unit_conversions
                                  (property_id, item_id, from_unit, to_unit, factor)
                                VALUES (%s, %s, %s, %s, %s)
                                """,
                                (PROPERTY_ID, iid, "box", "pieces", pieces_per),
                            )
                            cur.execute(
                                """
                                UPDATE stock_items
                                SET purchase_unit = COALESCE(purchase_unit, 'box'),
                                    usage_unit = COALESCE(usage_unit, 'pieces'),
                                    conversion_factor = %s
                                WHERE id = %s
                                """,
                                (pieces_per, iid),
                            )
                            count += 1
            continue
        # Section 1 fixed conversions (rows with item name in col 0)
        if not c0 or str(c0).strip().startswith(" ") and "SECTION" in str(c0).upper():
            continue
        if str(c0).lower().startswith("item name"):
            continue
        if "SECTION" in str(c0).upper() or str(c0).startswith("Select item"):
            continue
        item_name = str(c0).strip()
        from_unit = row[1] if len(row) > 1 else None
        factor = to_decimal(row[6]) if len(row) > 6 else Decimal("0")
        to_unit = row[8] if len(row) > 8 else None
        if not from_unit or not to_unit or factor <= 0:
            continue
        iid = resolve_item(by_sku, by_name, None, item_name)
        if not iid:
            issues.append(f"Conversion: unknown item {item_name!r}")
            continue
        fu, tu = str(from_unit).strip()[:40], str(to_unit).strip()[:40]
        key = (iid, norm_key(fu), norm_key(tu))
        if key in seen:
            continue
        seen.add(key)
        cur.execute(
            """
            INSERT INTO stock_unit_conversions
              (property_id, item_id, from_unit, to_unit, factor)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (PROPERTY_ID, iid, fu, tu, factor),
        )
        cur.execute(
            """
            UPDATE stock_items
            SET purchase_unit = %s, usage_unit = %s, conversion_factor = %s
            WHERE id = %s
            """,
            (fu, tu, factor, iid),
        )
        count += 1
    return count


def main() -> int:
    if not EXCEL_PATH.is_file():
        log(f"Excel not found: {EXCEL_PATH}")
        return 1

    log(f"Loading {EXCEL_PATH.name}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    conn = pymysql.connect(**DB, autocommit=False)
    try:
        with conn.cursor() as cur:
            log("Clearing existing property 1 warehouse data...")
            clear_property_stock(cur)

            log("Importing master items...")
            items, _wh = import_master(wb, cur)
            by_sku, by_name = build_item_lookup(
                {iid: {"sku": d["sku"], "name": d["name"]} for iid, d in items.items()}
            )

            log("Ensuring items from location-only sheets...")
            ensure_items_from_location_sheet(
                wb["Front Office Stock"], cur, by_sku, by_name, "front_office"
            )

            log("Applying location balances from sheets...")
            apply_location_sheet(
                wb["Kitchen Stock"], cur, by_sku, by_name, LOC["kitchen"], 6, "Kitchen Stock"
            )
            apply_location_sheet(
                wb["Cleaners Stock"], cur, by_sku, by_name, LOC["cleaners"], 6, "Cleaners Stock"
            )
            apply_location_sheet(
                wb["Front Office Stock"], cur, by_sku, by_name, LOC["front_office"], 6, "Front Office Stock"
            )

            log("Syncing quantity_on_hand from balances...")
            sync_quantity_on_hand(cur, items)

            log("Importing purchases (history only)...")
            n_purchases, n_purchase_lines = import_purchases(
                wb["Purchases"], cur, by_sku, by_name
            )

            log("Importing transfers (history only)...")
            n_transfers, n_transfer_lines = import_transfers(
                wb["All Transfers"], cur, by_sku, by_name
            )

            log("Importing conversions...")
            n_conversions = import_conversions(wb, cur, by_sku, by_name)

            conn.commit()

            cur.execute("SELECT COUNT(*) FROM stock_items WHERE property_id = %s", (PROPERTY_ID,))
            n_items = cur.fetchone()[0]
            cur.execute(
                """
                SELECT COUNT(*) FROM stock_balances sb
                JOIN stock_items si ON si.id = sb.item_id
                WHERE si.property_id = %s
                """,
                (PROPERTY_ID,),
            )
            n_balances = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM stock_purchases WHERE property_id = %s", (PROPERTY_ID,))
            n_purch = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM stock_transfers WHERE property_id = %s", (PROPERTY_ID,))
            n_trf = cur.fetchone()[0]
            cur.execute(
                "SELECT COUNT(*) FROM stock_unit_conversions WHERE property_id = %s", (PROPERTY_ID,)
            )
            n_conv = cur.fetchone()[0]

        log("\n=== Import counts ===")
        log(f"stock_items: {n_items}")
        log(f"stock_balances: {n_balances}")
        log(f"stock_purchases: {n_purch}")
        log(f"stock_transfer headers: {n_trf}")
        log(f"stock_unit_conversions: {n_conversions}")
        log(f"(purchase lines inserted: {n_purchase_lines}, transfer lines: {n_transfer_lines})")

        if issues:
            log(f"\n=== Issues ({len(issues)}) — showing up to 30 ===")
            for msg in issues[:30]:
                log(f"  - {msg}")
            if len(issues) > 30:
                log(f"  ... and {len(issues) - 30} more")
        else:
            log("\nNo issues recorded.")

    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
        wb.close()

    return 0


if __name__ == "__main__":
    sys.exit(main())
